#!/usr/bin/env python3
"""
Estrae i lead dal Google Sheet e li salva come CSV sul Desktop.
Esegui con: python3 ~/Desktop/estrai_lead.py
"""
import json, base64, io, csv, sys
from pathlib import Path

# Il file xlsx scaricato da Claude (path fisso per questa sessione)
CACHE_FILE = "/var/folders/5g/3llfgxjd6rx6n0m994ycbr_r0000gn/T/claude-hostloop-plugins/d10871c981f4fc5f/projects/-Users-tonyvalentinogallitto-Library-Application-Support-Claude-local-agent-mode-sessions-8b52f59f-73ae-43d1-a70b-0b08e74cfb89-65752a00-2d01-4f30-b141-20e405ed4ac5-local-4a3b5c28-54ef-4d04-8308-f23752-l8gijw/49c60d48-724e-43c0-9a94-2a47ca7e2c5b/tool-results/mcp-03410b0c-1679-4901-9edb-dc670df6787b-download_file_content-1779096214379.txt"
OUTPUT_CSV = str(Path.home() / "Desktop" / "leads.csv")

TARGET_CATEGORIES = {"Ottimale", "Alto"}

def main():
    try:
        import openpyxl
    except ImportError:
        print("Installa openpyxl: pip3 install openpyxl")
        sys.exit(1)

    print("📂 Leggo il file...")
    with open(CACHE_FILE) as f:
        data = json.load(f)

    xlsx_bytes = base64.b64decode(data["content"])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    print(f"📋 Fogli trovati: {wb.sheetnames}")

    all_leads = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

        def col(row, name):
            try:
                return str(row[headers.index(name)].value or "").strip()
            except (ValueError, IndexError):
                return ""

        for row in ws.iter_rows(min_row=2):
            categoria = col(row, "Categoria")
            if categoria in TARGET_CATEGORIES:
                all_leads.append({
                    "sheet":        sheet_name,
                    "categoria":    categoria,
                    "score":        col(row, "Score"),
                    "nome_azienda": col(row, "Nome Azienda"),
                    "telefono":     col(row, "Telefono"),
                    "regione":      col(row, "Regione"),
                    "citta":        col(row, "Città"),
                    "sito_web":     col(row, "Sito Web"),
                    "instagram":    col(row, "Instagram"),
                    "linkedin":     col(row, "LinkedIn"),
                    "email":        "",  # da arricchire
                })

    print(f"✅ Lead trovati (Ottimale + Alto): {len(all_leads)}")

    with open(OUTPUT_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_leads[0].keys())
        writer.writeheader()
        writer.writerows(all_leads)

    print(f"💾 Salvato in: {OUTPUT_CSV}")

    # Riepilogo per foglio
    from collections import Counter
    by_sheet = Counter(l["sheet"] for l in all_leads)
    for sheet, count in sorted(by_sheet.items()):
        print(f"   {sheet}: {count} lead")

if __name__ == "__main__":
    main()
